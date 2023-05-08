import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay, classification_report
import openpyxl
import xlsxwriter
from pathlib import Path
import xlsxwriter.utility as xl_utility

def plot_confusion_signals(X_test, y_test, y_pred, all_labels, output_path, n_split=0):
    fig, ax = plt.subplots(len(all_labels), len(all_labels), figsize=(len(all_labels)*5, len(all_labels)*5))

    for idx_pred, label_pred in enumerate(all_labels):
        for idx_true, label_true in enumerate(all_labels):
            bool = (y_test == idx_true) & (y_pred == idx_pred)

            if sum(bool) > 0:
                if label_pred == label_true:
                    ax[idx_true, idx_pred].plot(X_test[bool].T, c="g", alpha=0.3)
                else:
                    ax[idx_true, idx_pred].plot(X_test[bool].T, c="r", alpha=0.3)

            ax[idx_true, idx_pred].set_ylim((X_test.min(), X_test.max()))
            ax[idx_true, idx_pred].set_title(f"{label_true} classified {label_pred} ({sum(bool)})", fontsize=10)

    plt.tight_layout()
    plt.savefig(output_path / f"{n_split}_confusion_signals")


def plot_confusion_matrix(y_test_argmax, y_pred_argmax, all_labels, output_path, n_split=0):
    # Plot confusion matrix
    result = classification_report(y_test_argmax, y_pred_argmax, target_names=all_labels)

    fig, ax = plt.subplots(1, 2, figsize=(15, 8))
    cm = confusion_matrix(y_test_argmax, y_pred_argmax)
    disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=all_labels).plot(ax=ax[0])
    ax[0].set_title(f"Split {n_split}")
    ax[0].set_xticklabels(all_labels, rotation=45)
    ax[1].text(0, 1,
               result,
               horizontalalignment='left',
               verticalalignment='top')
    ax[1].set_axis_off()
    plt.tight_layout()
    plt.savefig(output_path / f"{n_split}_confusion_matrix")


def write_excel(df_result_magnet: pd.DataFrame, labels_path: Path, neighbor_labelevent: pd.DataFrame,  output_path:Path):
    """
    function stores result of df_result_magnet in Excel
    :param df_result_magnet: Dataframe with magnets in index and labels as columns
    :param labels_path: path where input excel with labels is stored
    :param output_path: path where output file should be stored
    """
    # Open the source Excel file
    source_workbook = openpyxl.load_workbook(labels_path)
    source_sheet = source_workbook.active
    df_event_labels = pd.read_excel(labels_path)

    #with open(output_path / 'dest_file.xlsx', 'wb') as dest_file:
    dest_workbook = xlsxwriter.Workbook(output_path)
    dest_sheet = dest_workbook.add_worksheet()

    url_links = [source_sheet.cell(row=2 + i, column=2).hyperlink.target for i in range(len(df_event_labels))]
    df_event_labels["url"] = url_links

    columns = ['event', 'Electrical order'] \
              + list(df_result_magnet.columns) \
              + [f"NN_{c}" for c in df_result_magnet.filter(regex="pred").columns] \
              + neighbor_labelevent.filter(regex="_kneighbors").columns.to_list()
    dest_sheet.write_row(0, 0, columns)

    red_format = dest_workbook.add_format({'bg_color': '#FFC7CE'})
    i = 1
    for event, row in df_result_magnet.iterrows():
        # Read the URL from the source file
        index = df_event_labels[df_event_labels.event == event].index.values[0]

        url = source_sheet.cell(row=2 + index, column=2).hyperlink.target
        el_order = source_sheet.cell(row=2 + index, column=2).value

        if url != df_event_labels[df_event_labels.event == event].url.values[0]:
            print("error")

        # Create a new Excel file and copy the URL to it
        dest_sheet.write(i, 0, event)
        dest_sheet.write_url(i, 1, url, string=str(el_order))
        for j, entry in enumerate(row.index.values):
            dest_sheet.write(i, j + 2, row[entry])

        if neighbor_labelevent is not None:
            first_empty_column = len(row) + 2
            kneighbors = neighbor_labelevent.filter(regex="_kneighbors").loc[event].values
            distance = neighbor_labelevent.filter(regex="_distance").loc[event].values
            t = 0
            for k, d in zip(kneighbors, distance):
                url = df_event_labels["url"].iloc[int(k)]
                event = df_event_labels["event"].iloc[int(k)]
                dest_sheet.write_url(i, first_empty_column + t, url, string=str(d)) #d
                t += 1

            #for nl, neighbor_label in enumerate(neighbor_labelevent[i-1]):  # iterate over labels
            #        dest_sheet.write(i, first_empty_column + nl,  str(neighbor_label))
            #    else:
            #        dest_sheet.write(i, first_empty_column + nl,  0)
        i += 1

    for column_index in range(len(columns)):
        column_name = xl_utility.xl_col_to_name(column_index)
        dest_sheet.set_column(f'{column_name}:{column_name}', 20)

    dest_workbook.close()

class WindowSlice():
    def __init__(self, labels=None, reduce_ratio=0.9, none_flat_shape=None):
        self.labels = labels
        self.reduce_ratio = reduce_ratio
        self.none_flat_shape = none_flat_shape
        self.name = "window_slice"

    def augment(self, X: np.array, y: np.array, oversampling_rate=None) -> tuple:
        encoded = False
        if y.ndim == 1: # label has to be one hot encoded
            encoded = True
            y = np.identity(2)[y]

        index = np.arange(len(X))
        biggest_label_size = np.sum(y, axis=0).max()
        if oversampling_rate is None:
            oversampling_rate = biggest_label_size / np.sum(y, axis=0) - 1  # ratio of artificial labels

        shape_X = X.shape
        if self.none_flat_shape:
            X = X.reshape(self.none_flat_shape)
        else:
            X = X.reshape((X.shape[0], -1, X.shape[-1]))

        X_augmented_list = []
        y_augmented_list = []
        source_index_list = []
        for i, label in enumerate(self.labels):
            label_bool = (y[:, i] == label[i])

            X_label = X[label_bool]
            y_label = y[label_bool]

            if len(X_label) == biggest_label_size:
                random_choices = np.random.choice(len(X_label), size=int(oversampling_rate[i] * len(X_label)), replace=False)
            else:
                random_choices = np.random.choice(len(X_label), size=int(oversampling_rate[i] * len(X_label)))

            X_oversampled = X_label[random_choices]
            y_oversampled = y_label[random_choices]

            # https://halshs.archives-ouvertes.fr/halshs-01357973/document
            X_window_sliced = np.zeros_like(X_oversampled)
            target_len = np.ceil(self.reduce_ratio * X_oversampled.shape[-1]).astype(int)

            starts = np.random.randint(low=0, high=X.shape[-1] - target_len, size=(X_oversampled.shape[0])).astype(int)
            ends = (target_len + starts).astype(int)

            for j, pattern in enumerate(X_oversampled):
                for dim in range(X_oversampled.shape[1]):
                    X_window_sliced[j, dim, :] = np.interp(np.linspace(0, target_len, num=X_oversampled.shape[-1]),
                                                           np.arange(target_len),
                                                           pattern[dim, starts[j]:ends[j]]).T

            # combine augmented and original labels
            X_augm_label = np.vstack((X_label.reshape(X_label.shape[0], -1),
                                     X_window_sliced.reshape((-1, ) + shape_X[1:])))
            y_augm_label = np.vstack((y_label, y_oversampled))
            idx_augm_label = np.hstack((index[label_bool], index[label_bool][random_choices]))

            X_augmented_list.append(X_augm_label)
            y_augmented_list.append(y_augm_label)
            source_index_list.append(idx_augm_label)

        X_augmented = np.vstack(X_augmented_list)
        y_augmented = np.vstack(y_augmented_list)
        source_index = np.hstack(source_index_list)
        #print(f"len: {len(X)}, unique: {pd.DataFrame(source_index).nunique().values}")

        if encoded == True:
            y_augmented = np.argmax(y_augmented, axis=1)

        idx_shuffled = np.random.permutation(np.arange(len(X_augmented)))

        return X_augmented[idx_shuffled], y_augmented[idx_shuffled], source_index[idx_shuffled]




