import pandas as pd
from matplotlib import pyplot as plt
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay, classification_report
import openpyxl
import xlsxwriter
from pathlib import Path
import xlsxwriter.utility as xl_utility

def plot_confusion_signals(X_test, y_test, y_pred, all_labels, output_path, n_split=0):
    fig, ax = plt.subplots(len(all_labels), len(all_labels), figsize=(len(all_labels)*5, len(all_labels)*5))

    for idx_pred, label_pred in enumerate(all_labels):
        for idx_true, label_true in enumerate(all_labels):
            bool = (y_test == idx_true) & (y_pred == idx_pred)

            if sum(bool) > 0:
                if label_pred == label_true:
                    ax[idx_true, idx_pred].plot(X_test[bool].T, c="g", alpha=0.3)
                else:
                    ax[idx_true, idx_pred].plot(X_test[bool].T, c="r", alpha=0.3)

            ax[idx_true, idx_pred].set_ylim((X_test.min(), X_test.max()))
            ax[idx_true, idx_pred].set_title(f"{label_true} classified {label_pred} ({sum(bool)})", fontsize=10)

    plt.tight_layout()
    plt.savefig(output_path / f"{n_split}_confusion_signals")


def plot_confusion_matrix(y_test_argmax, y_pred_argmax, all_labels, output_path, n_split=0):
    # Plot confusion matrix
    result = classification_report(y_test_argmax, y_pred_argmax, target_names=all_labels)

    fig, ax = plt.subplots(1, 2, figsize=(15, 8))
    cm = confusion_matrix(y_test_argmax, y_pred_argmax)
    disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=all_labels).plot(ax=ax[0])
    ax[0].set_title(f"Split {n_split}")
    ax[0].set_xticklabels(all_labels, rotation=45)
    ax[1].text(0, 1,
               result,
               horizontalalignment='left',
               verticalalignment='top')
    ax[1].set_axis_off()
    plt.tight_layout()
    plt.savefig(output_path / f"{n_split}_confusion_matrix")


def write_excel(df_result_magnet: pd.DataFrame, labels_path:Path, neighbor_labelevent,  output_path:Path):
    """
    function stores result of df_result_magnet in Excel
    :param df_result_magnet: Dataframe with magnets in index and labels as columns
    :param labels_path: path where input excel with labels is stored
    :param output_path: path where output file should be stored
    """
    # Open the source Excel file
    source_workbook = openpyxl.load_workbook(labels_path)
    source_sheet = source_workbook.active
    df_event_labels = pd.read_excel(labels_path)

    #with open(output_path / 'dest_file.xlsx', 'wb') as dest_file:
    dest_workbook = xlsxwriter.Workbook(output_path / 'dest_file.xlsx')
    dest_sheet = dest_workbook.add_worksheet()

    columns = ['event', 'Electrical order'] \
              + list(df_result_magnet.columns) \
              + [f"NN_{c}" for c in df_result_magnet.filter(regex="pred").columns]
    dest_sheet.write_row(0, 0, columns)

    red_format = dest_workbook.add_format({'bg_color': '#FFC7CE'})
    i = 1
    for event, row in df_result_magnet.iterrows():
        # Read the URL from the source file
        index = df_event_labels[df_event_labels.event == event].index.values[0]

        url = source_sheet.cell(row=2 + index, column=2).hyperlink.target
        el_order = source_sheet.cell(row=2 + index, column=2).value

        # Create a new Excel file and copy the URL to it
        dest_sheet.write(i, 0, event)
        dest_sheet.write_url(i, 1, url, string=str(el_order))
        for j, entry in enumerate(row.index.values):
            dest_sheet.write(i, j + 2, row[entry])

        first_empty_column = len(row) + 2
        for nl, neighbor_label in enumerate(neighbor_labelevent[i-1]): # iterate over labels
            if neighbor_label != []:
                dest_sheet.write(i, first_empty_column + nl,  str(neighbor_label))
            else:
                dest_sheet.write(i, first_empty_column + nl,  0)
        i += 1

    for column_index in range(len(columns)):
        column_name = xl_utility.xl_col_to_name(column_index)
        dest_sheet.set_column(f'{column_name}:{column_name}', 20)

    dest_workbook.close()